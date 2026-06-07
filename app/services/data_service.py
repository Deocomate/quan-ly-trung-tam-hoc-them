# app/services/data_service.py
from __future__ import annotations
import csv
import io
import sqlite3
import zipfile
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select, text
from sqlalchemy.orm import Session
from app.database import engine
from app.models import Student, Class, Enrollment, Attendance, TuitionRecord, TuitionRecordItem, Setting, User

# Bản đồ tên bảng tương ứng với phân nhóm cấu hình
TABLE_MAP = {
    "students": ["enrollments", "students"],
    "classes": ["classes"],
    "attendance": ["attendance"],
    "tuition": ["tuition_record_items", "tuition_records", "tuition_periods"],
    "settings": ["settings"],
    "users": ["users"],
    "teachers": ["teachers"],
    "payroll": ["teacher_salary_record_items", "teacher_salary_records"]
}

# 1. LOGIC XUẤT SQL DUMP SỬ DỤNG SQLITE3 ITERDUMP HOẶC CHỌN LỌC
def export_to_sql(targets: list[str]) -> str:
    """
    Sinh mã SQL phục vụ sao lưu dựa trên các phân nhóm được chọn.
    """
    conn = engine.raw_connection()
    try:
        cursor = conn.cursor()
        sql_lines = []
        sql_lines.append("-- backup-school-management-system")
        sql_lines.append(f"-- Generated at: {datetime.now().isoformat()}\n")
        sql_lines.append("PRAGMA foreign_keys=OFF;")
        
        # Xác định danh sách bảng cụ thể cần backup dựa vào cấu hình target
        tables_to_dump = []
        for target in targets:
            tables_to_dump.extend(TABLE_MAP.get(target, []))
            
        for table in tables_to_dump:
            # Drop table câu lệnh
            sql_lines.append(f"DROP TABLE IF EXISTS {table};")
            
            # Lấy cấu trúc CREATE TABLE ban đầu từ sqlite_master
            cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (table,))
            create_stmt = cursor.fetchone()
            if create_stmt and create_stmt[0]:
                sql_lines.append(create_stmt[0] + ";")
            
            # Lấy toàn bộ bản ghi để sinh INSERT INTO
            cursor.execute(f"SELECT * FROM {table}")
            rows = cursor.fetchall()
            if rows:
                # Lấy tên các cột
                col_names = [description[0] for description in cursor.description]
                col_str = ", ".join([f'"{name}"' for name in col_names])
                
                for row in rows:
                    vals = []
                    for val in row:
                        if val is None:
                            vals.append("NULL")
                        elif isinstance(val, (int, float)):
                            vals.append(str(val))
                        else:
                            # Escape ký tự nháy đơn trong chuỗi
                            escaped_val = str(val).replace("'", "''")
                            vals.append(f"'{escaped_val}'")
                    val_str = ", ".join(vals)
                    sql_lines.append(f"INSERT INTO {table} ({col_str}) VALUES ({val_str});")
                    
        sql_lines.append("PRAGMA foreign_keys=ON;")
        return "\n".join(sql_lines)
    finally:
        conn.close()

# 2. LOGIC PHỤC HỒI TỪ SQL (IMPORT)
def import_from_sql(sql_content: str) -> int:
    """
    Thực thi chuỗi lệnh SQL trong một transaction duy nhất để đảm bảo tính an toàn dữ liệu.
    """
    conn = engine.raw_connection()
    try:
        cursor = conn.cursor()
        # Bọc toàn bộ script trong một transaction để đảm bảo tính nguyên tử
        full_script = f"BEGIN TRANSACTION;\n{sql_content}\nCOMMIT;"
        # Chạy dưới dạng tập lệnh (Script) để hỗ trợ nhiều câu lệnh đồng thời
        cursor.executescript(full_script)
        # Đếm số lượng câu lệnh insert được thực hiện một cách tương đối
        statements_count = len([line for line in sql_content.splitlines() if line.strip().upper().startswith("INSERT")])
        return statements_count
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        raise exc
    finally:
        conn.close()

# 3. ĐỊNH NGHĨA HEADER VÀ DỮ LIỆU ĐỂ XUẤT EXCEL/CSV THÂN THIỆN
def get_dataset_structure(db: Session, targets: list[str]) -> dict[str, dict[str, Any]]:
    """
    Hàm chuẩn bị dữ liệu thô, ánh xạ tiêu đề tiếng Việt và định dạng cột dễ hiểu cho người dùng.
    """
    dataset = {}
    
    if "students" in targets:
        students = db.scalars(select(Student).order_by(Student.student_code)).all()
        dataset["Học sinh"] = {
            "headers": ["Mã học sinh", "Họ tên", "Số điện thoại phụ huynh", "Ghi chú", "Trạng thái hoạt động", "Ngày tạo"],
            "rows": [
                [
                    s.student_code,
                    s.full_name,
                    s.parent_phone or "",
                    s.notes or "",
                    "Đang học" if s.is_active else "Đã nghỉ",
                    s.created_at.strftime("%d/%m/%Y") if s.created_at else ""
                ]
                for s in students
            ]
        }
        
    if "classes" in targets:
        classes = db.scalars(select(Class).order_by(Class.name)).all()
        dataset["Lớp - Môn học"] = {
            "headers": ["Tên lớp", "Môn học", "Học phí mặc định (VNĐ)", "Ghi chú", "Trạng thái hoạt động"],
            "rows": [
                [c.name, c.subject, c.default_fee, c.notes or "", "Đang dạy" if c.is_active else "Ngưng hoạt động"]
                for c in classes
            ]
        }
        
    if "attendance" in targets:
        # Lấy dữ liệu điểm danh kết hợp thông tin học sinh và lớp
        stmt = select(Attendance).join(Attendance.student).join(Attendance.class_).order_by(Attendance.date.desc(), Student.full_name)
        attendances = db.scalars(stmt).all()
        status_translation = {"P": "Có mặt", "V": "Vắng", "M": "Muộn"}
        dataset["Điểm danh chi tiết"] = {
            "headers": ["Ngày học", "Mã học sinh", "Tên học sinh", "Lớp học", "Trạng thái"],
            "rows": [
                [
                    att.date.strftime("%d/%m/%Y") if att.date else "",
                    att.student.student_code,
                    att.student.full_name,
                    att.class_.name,
                    status_translation.get(att.status, att.status)
                ]
                for att in attendances
            ]
        }
        
    if "tuition" in targets:
        records = db.scalars(select(TuitionRecord).order_by(TuitionRecord.year.desc(), TuitionRecord.month.desc())).all()
        dataset["Báo cáo Học phí"] = {
            "headers": ["Kỳ học phí", "Mã học sinh", "Tên học sinh", "Tổng số buổi", "Tổng tiền phải nộp (VNĐ)"],
            "rows": [
                [f"Tháng {r.month}/{r.year}", r.student.student_code, r.student.full_name, r.total_sessions, r.total_amount]
                for r in records
            ]
        }
        
    if "settings" in targets:
        settings = db.scalars(select(Setting).order_by(Setting.key)).all()
        dataset["Cấu hình trung tâm"] = {
            "headers": ["Khóa cấu hình", "Nội dung thiết lập"],
            "rows": [[s.key, s.value] for s in settings]
        }
        
    if "users" in targets:
        users = db.scalars(select(User).order_by(User.username)).all()
        dataset["Tài khoản quản trị"] = {
            "headers": ["Tên đăng nhập", "Họ tên người dùng", "Trạng thái hoạt động"],
            "rows": [[u.username, u.full_name, "Đang hoạt động" if u.is_active else "Khóa"] for u in users]
        }
        
    if "teachers" in targets:
        from app.models import Teacher
        teachers = db.scalars(select(Teacher).order_by(Teacher.full_name)).all()
        dataset["Giáo viên"] = {
            "headers": ["Họ tên", "Số điện thoại", "Email", "Hệ số lương mặc định", "Trạng thái"],
            "rows": [
                [t.full_name, t.phone or "", t.email or "", t.default_salary_coefficient, "Đang làm việc" if t.is_active else "Đã nghỉ"]
                for t in teachers
            ]
        }
        
    if "payroll" in targets:
        from app.models import TeacherSalaryRecord
        records = db.scalars(select(TeacherSalaryRecord).order_by(TeacherSalaryRecord.year.desc(), TeacherSalaryRecord.month.desc())).all()
        dataset["Lịch sử lương giáo viên"] = {
            "headers": ["Kỳ lương", "Tên giáo viên", "Tổng lương (VNĐ)", "Trạng thái", "Ngày chốt"],
            "rows": [
                [
                    f"Tháng {r.month}/{r.year}",
                    r.teacher.full_name,
                    r.total_amount,
                    "Đã chốt" if r.is_locked else "Tạm tính",
                    r.locked_at.strftime("%d/%m/%Y") if r.locked_at else ""
                ]
                for r in records
            ]
        }
        
    return dataset

# 4. TẠO WORKBOOK EXCEL TRỰC QUAN NHIỀU SHEETS
def generate_formatted_excel(db: Session, targets: list[str]) -> bytes:
    dataset = get_dataset_structure(db, targets)
    wb = Workbook()
    
    # Loại bỏ sheet mặc định ban đầu
    default_sheet = wb.active
    if default_sheet:
        wb.remove(default_sheet)
        
    # Kiểu dáng định dạng chung
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="10202B")
    fill_header = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="CBD5DC"),
        right=Side(style="thin", color="CBD5DC"),
        top=Side(style="thin", color="CBD5DC"),
        bottom=Side(style="thin", color="CBD5DC")
    )
    
    for sheet_name, content in dataset.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        # Tiêu đề bảng lớn ở dòng 1
        ws.cell(row=1, column=1, value=sheet_name.upper()).font = Font(name=font_family, size=14, bold=True, color="0F766E")
        ws.row_dimensions[1].height = 28
        
        # Thiết lập header ở dòng 3
        ws.row_dimensions[3].height = 24
        for col_idx, h_text in enumerate(content["headers"], start=1):
            cell = ws.cell(row=3, column=col_idx, value=h_text)
            cell.font = header_font
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Thêm các dòng dữ liệu
        for row_idx, row_data in enumerate(content["rows"], start=4):
            ws.row_dimensions[row_idx].height = 20
            row_fill = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                if row_fill.fill_type:
                    cell.fill = row_fill
                
                # Định dạng số đối với cột tiền tệ hoặc số lượng
                if isinstance(value, int) and value > 1000:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif isinstance(value, int):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        # Tự động căn chỉnh độ rộng cột vừa vặn
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue  # Bỏ qua dòng tiêu đề lớn khi đo kích thước cột
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

# 5. TẠO TỆP ZIP CHỨA CÁC TỆP CSV RIÊNG BIỆT
def generate_csv_zip(db: Session, targets: list[str]) -> bytes:
    dataset = get_dataset_structure(db, targets)
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for sheet_name, content in dataset.items():
            csv_buffer = StringIO()
            # Sử dụng utf-8-sig để Excel hiển thị tiếng Việt mà không lỗi font (BOM)
            writer = csv.writer(csv_buffer)
            writer.writerow(content["headers"])
            writer.writerows(content["rows"])
            
            file_name = f"{sheet_name.lower().replace(' ', '_').replace('-', '_')}.csv"
            zip_file.writestr(file_name, csv_buffer.getvalue().encode("utf-8-sig"))
            
    return zip_buffer.getvalue()
