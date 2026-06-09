from io import BytesIO
import datetime
from sqlalchemy import func, select, distinct
from sqlalchemy.orm import Session, selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import TuitionRecord, TuitionRecordItem, Student, Class, Enrollment


def format_period_label(months: list[int], year: int) -> tuple[str, str]:
    if len(months) == 1:
        m = months[0]
        return f"THÁNG {m:02d}/{year}", f"T{m:02d}-{year}"
    elif len(months) == 3:
        sorted_m = sorted(months)
        if sorted_m == [1, 2, 3]:
            q = 1
        elif sorted_m == [4, 5, 6]:
            q = 2
        elif sorted_m == [7, 8, 9]:
            q = 3
        else:
            q = 4
        return f"QUÝ {q}/{year}", f"Quy {q}-{year}"
    elif len(months) == 12:
        return f"NĂM {year}", f"{year}"
    else:
        sorted_m = sorted(months)
        m_range = f"{sorted_m[0]:02d}-{sorted_m[-1]:02d}"
        return f"KỲ {m_range}/{year}", f"Ky {m_range}-{year}"


def generate_revenue_report_excel(db: Session, month: int | list[int], year: int, settings: dict[str, str]) -> bytes:
    if isinstance(month, int):
        months = [month]
    else:
        months = month

    # 1. Lấy dữ liệu tổng quan
    total_students = db.scalar(
        select(func.count(distinct(TuitionRecord.student_id)))
        .where(TuitionRecord.month.in_(months), TuitionRecord.year == year)
    ) or 0

    total_classes = db.scalar(
        select(func.count(distinct(TuitionRecordItem.class_id)))
        .join(TuitionRecordItem.record)
        .where(TuitionRecord.month.in_(months), TuitionRecord.year == year)
    ) or 0

    total_sessions = db.scalar(
        select(func.coalesce(func.sum(TuitionRecordItem.sessions), 0))
        .join(TuitionRecordItem.record)
        .where(TuitionRecord.month.in_(months), TuitionRecord.year == year)
    ) or 0

    total_revenue = db.scalar(
        select(func.coalesce(func.sum(TuitionRecord.total_amount), 0))
        .where(TuitionRecord.month.in_(months), TuitionRecord.year == year)
    ) or 0

    class_rows = db.execute(
        select(
            TuitionRecordItem.class_name,
            TuitionRecordItem.subject,
            func.count(distinct(TuitionRecord.student_id)).label("students_count"),
            func.coalesce(func.sum(TuitionRecordItem.sessions), 0).label("sessions_count"),
            func.coalesce(func.sum(TuitionRecordItem.amount), 0).label("revenue")
        )
        .join(TuitionRecordItem.record)
        .where(TuitionRecord.month.in_(months), TuitionRecord.year == year)
        .group_by(TuitionRecordItem.class_id, TuitionRecordItem.class_name, TuitionRecordItem.subject)
        .order_by(TuitionRecordItem.class_name)
    ).all()

    # 2. Tạo workbook và sheet
    wb = Workbook()
    ws = wb.active
    period_title, sheet_title = format_period_label(months, year)
    ws.title = f"Doanh thu {sheet_title}"

    # Bật hiển thị đường lưới
    ws.views.sheetView[0].showGridLines = True

    # 3. Định nghĩa style và font
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="0F2A33")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="526672")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True, color="10202B")
    regular_font = Font(name=font_family, size=11, color="10202B")
    
    fill_header = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    fill_summary = PatternFill(start_color="E6F4F1", end_color="E6F4F1", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5DC"),
        right=Side(style="thin", color="CBD5DC"),
        top=Side(style="thin", color="CBD5DC"),
        bottom=Side(style="thin", color="CBD5DC")
    )
    double_bottom_border = Border(
        left=Side(style="thin", color="CBD5DC"),
        right=Side(style="thin", color="CBD5DC"),
        top=Side(style="thin", color="CBD5DC"),
        bottom=Side(style="double", color="10202B")
    )

    # 4. Phần đầu đề (Header báo cáo)
    center_name = settings.get("center_name", "TRUNG TÂM GIÁO DỤC")
    ws["A1"] = center_name.upper()
    ws["A1"].font = Font(name=font_family, size=11, bold=True, color="0F766E")
    
    ws["A2"] = f"BÁO CÁO DOANH THU {period_title}"
    ws["A2"].font = title_font
    
    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws["A3"] = f"Ngày xuất báo cáo: {now_str}"
    ws["A3"].font = subtitle_font
    
    ws.row_dimensions[2].height = 25

    # 5. Bảng Tóm tắt Tổng quan
    ws["A5"] = "TỔNG QUAN DOANH THU"
    ws["A5"].font = Font(name=font_family, size=12, bold=True, color="0F766E")
    
    summary_headers = ["Chỉ số", "Giá trị"]
    for col_idx, h in enumerate(summary_headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    summary_data = [
        ("Tổng số học sinh đã chốt học phí", total_students, "students"),
        ("Tổng số lớp học phát sinh học phí", total_classes, "classes"),
        ("Tổng số buổi học đã tham gia", total_sessions, "sessions"),
        ("Tổng doanh thu thực tế (VNĐ)", total_revenue, "revenue")
    ]
    
    for row_idx, (label, val, unit) in enumerate(summary_data, start=7):
        c1 = ws.cell(row=row_idx, column=1, value=label)
        c1.font = regular_font
        c1.border = thin_border
        c1.alignment = Alignment(horizontal="left", vertical="center")
        
        c2 = ws.cell(row=row_idx, column=2, value=val)
        c2.font = bold_font if unit == "revenue" else regular_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal="right", vertical="center")
        if unit == "revenue":
            c2.number_format = "#,##0"
            c2.fill = fill_summary
            
    # 6. Bảng Chi tiết theo Từng Lớp Học
    start_class_row = 13
    ws.cell(row=start_class_row - 1, column=1, value="CHI TIẾT DOANH THU THEO TỪNG LỚP HỌC").font = Font(name=font_family, size=12, bold=True, color="0F766E")
    
    class_headers = ["STT", "Tên lớp", "Môn học", "Sỹ số (HS)", "Tổng buổi học", "Doanh thu (VNĐ)", "Tỷ lệ (%)"]
    ws.row_dimensions[start_class_row].height = 25
    for col_idx, h in enumerate(class_headers, start=1):
        cell = ws.cell(row=start_class_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    current_row = start_class_row + 1
    for idx, row in enumerate(class_rows, start=1):
        ws.row_dimensions[current_row].height = 20
        row_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
        
        c_stt = ws.cell(row=current_row, column=1, value=idx)
        c_stt.alignment = Alignment(horizontal="center", vertical="center")
        
        c_name = ws.cell(row=current_row, column=2, value=row.class_name)
        c_name.alignment = Alignment(horizontal="left", vertical="center")
        
        c_subject = ws.cell(row=current_row, column=3, value=row.subject)
        c_subject.alignment = Alignment(horizontal="left", vertical="center")
        
        c_stud = ws.cell(row=current_row, column=4, value=row.students_count)
        c_stud.alignment = Alignment(horizontal="center", vertical="center")
        c_stud.number_format = "#,##0"
        
        c_sess = ws.cell(row=current_row, column=5, value=row.sessions_count)
        c_sess.alignment = Alignment(horizontal="center", vertical="center")
        c_sess.number_format = "#,##0"
        
        c_rev = ws.cell(row=current_row, column=6, value=row.revenue)
        c_rev.alignment = Alignment(horizontal="right", vertical="center")
        c_rev.number_format = "#,##0"
        c_rev.font = bold_font
        
        # Công thức tính tỷ lệ phần trăm so với tổng doanh thu nằm ở ô B10
        c_ratio = ws.cell(row=current_row, column=7, value=f"=F{current_row}/$B$10")
        c_ratio.alignment = Alignment(horizontal="right", vertical="center")
        c_ratio.number_format = "0.0%"
        
        for c in [c_stt, c_name, c_subject, c_stud, c_sess, c_rev, c_ratio]:
            c.font = regular_font if c != c_rev else bold_font
            c.border = thin_border
            if row_fill.fill_type:
                c.fill = row_fill
                
        current_row += 1

    # Dòng tổng cộng cho bảng chi tiết lớp học
    ws.row_dimensions[current_row].height = 22
    ws.cell(row=current_row, column=1, value="Tổng cộng").font = bold_font
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    
    c_tot_stud = ws.cell(row=current_row, column=4, value=f"=SUM(D{start_class_row+1}:D{current_row-1})")
    c_tot_stud.alignment = Alignment(horizontal="center", vertical="center")
    c_tot_stud.number_format = "#,##0"
    
    c_tot_sess = ws.cell(row=current_row, column=5, value=f"=SUM(E{start_class_row+1}:E{current_row-1})")
    c_tot_sess.alignment = Alignment(horizontal="center", vertical="center")
    c_tot_sess.number_format = "#,##0"
    
    c_tot_rev = ws.cell(row=current_row, column=6, value=f"=SUM(F{start_class_row+1}:F{current_row-1})")
    c_tot_rev.alignment = Alignment(horizontal="right", vertical="center")
    c_tot_rev.number_format = "#,##0"
    
    c_tot_ratio = ws.cell(row=current_row, column=7, value=f"=SUM(G{start_class_row+1}:G{current_row-1})")
    c_tot_ratio.alignment = Alignment(horizontal="right", vertical="center")
    c_tot_ratio.number_format = "0.0%"

    for col in range(1, 8):
        c = ws.cell(row=current_row, column=col)
        c.font = bold_font
        c.fill = fill_summary
        c.border = double_bottom_border

    # 7. Tự động co giãn kích thước cột vừa văn dữ liệu
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Bỏ qua đo độ dài dòng tiêu đề lớn ở cột A để tránh cột A quá rộng
            if col_letter == "A" and cell.row in [1, 2, 3, 5, 12]:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Lưu workbook vào BytesIO để trả về
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_class_attendance_excel(
    db: Session,
    class_id: int,
    month: int,
    year: int,
    settings: dict[str, str],
    session_days_str: str | None = None,
    fill_attendance: bool = False
) -> bytes:
    import datetime
    from app.models import Class, Student, Enrollment, Attendance
    
    cls = db.get(Class, class_id)
    if not cls:
        raise ValueError("Class not found")
        
    # 1. Lấy danh sách học sinh thuộc lớp
    stmt = (
        select(Student)
        .join(Enrollment, Enrollment.student_id == Student.id)
        .where(
            Enrollment.class_id == class_id,
            Enrollment.is_active == True,
            Student.is_active == True
        )
    )
    students = db.scalars(stmt).all()
    
    # Sắp xếp theo thứ tự A-Z Tiếng Việt
    from app.routers.students import get_vietnamese_name_sort_key
    sorted_students = sorted(students, key=lambda s: get_vietnamese_name_sort_key(s.full_name))
    
    # 2. Xử lý danh sách ngày học
    session_dates = []
    if session_days_str and session_days_str.strip():
        parts = [p.strip() for p in session_days_str.split(",")]
        for p in parts:
            try:
                day_val = int(p)
                d = datetime.date(year, month, day_val)
                session_dates.append(d)
            except Exception:
                continue
    else:
        # Tự phát hiện từ dữ liệu điểm danh thực tế
        stmt_dates = (
            select(distinct(Attendance.date))
            .where(
                Attendance.class_id == class_id,
                func.strftime("%m", Attendance.date) == f"{month:02d}",
                func.strftime("%Y", Attendance.date) == str(year)
            )
        )
        dates_found = db.scalars(stmt_dates).all()
        session_dates = sorted(dates_found)
        
    is_blank_sheet = len(session_dates) == 0
    num_cols = 10 if is_blank_sheet else len(session_dates)
    total_cols = 4 + num_cols + 1
    col_letter_last = get_column_letter(total_cols)

    # 3. Khởi tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Chuyen Can"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Times New Roman"
    
    title_style = Font(name=font_family, size=16, bold=True)
    subtitle_style = Font(name=font_family, size=11, italic=True)
    bold_style = Font(name=font_family, size=11, bold=True)
    regular_style = Font(name=font_family, size=11)
    
    border_side = Side(style="thin", color="000000")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    parent_org = (settings.get("center_parent_org") or "").strip()
    ws["A1"] = parent_org.upper() if parent_org else ""
    ws["A1"].font = bold_style
    ws["A2"] = settings.get("center_name", "TRUNG TÂM GIÁO DỤC").upper()
    ws["A2"].font = bold_style

    ws["A3"] = f"DANH SÁCH HỌC SINH LỚP {cls.name.upper()}"
    ws["A3"].font = title_style
    ws["A3"].alignment = center_align

    ws["A4"] = f"năm học {cls.school_year or '2025-2026'}"
    ws["A4"].font = subtitle_style
    ws["A4"].alignment = center_align

    ws.merge_cells(f"A3:{col_letter_last}3")
    ws.merge_cells(f"A4:{col_letter_last}4")
    
    ws["A5"] = f"Tháng {month}/Q H"
    ws["A5"].font = bold_style
    
    cell_mon = ws.cell(row=5, column=total_cols - 1)
    cell_mon.value = f"Môn: {cls.subject.upper()}"
    cell_mon.font = Font(name=font_family, size=12, bold=True)
    cell_mon.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=5, start_column=total_cols-2, end_row=5, end_column=total_cols)
    
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 20
    
    ws.row_dimensions[6].height = 25
    
    headers = [("A6", "TT"), ("B6", "Mã học sinh"), ("C6", "Họ & tên"), ("D6", "Ngày sinh")]
    for cell_ref, val in headers:
        ws[cell_ref] = val
        ws[cell_ref].font = bold_style
        ws[cell_ref].alignment = center_align
        ws[cell_ref].border = thin_border

    for i in range(num_cols):
        col_idx = 5 + i
        cell = ws.cell(row=6, column=col_idx)
        cell.font = bold_style
        cell.alignment = center_align
        cell.border = thin_border
        if not is_blank_sheet:
            d = session_dates[i]
            cell.value = f"{d.day}/{d.month}"
        else:
            cell.value = ""
            
    cell_gc = ws.cell(row=6, column=total_cols)
    cell_gc.value = "Ghi chú"
    cell_gc.font = bold_style
    cell_gc.alignment = center_align
    cell_gc.border = thin_border
    
    current_row = 7
    for idx, student in enumerate(sorted_students, start=1):
        ws.row_dimensions[current_row].height = 20
        
        c_stt = ws.cell(row=current_row, column=1, value=idx)
        c_stt.font = regular_style
        c_stt.alignment = center_align
        c_stt.border = thin_border

        c_code = ws.cell(row=current_row, column=2, value=student.student_code)
        c_code.font = regular_style
        c_code.alignment = center_align
        c_code.border = thin_border

        c_name = ws.cell(row=current_row, column=3, value=student.full_name)
        c_name.font = regular_style
        c_name.alignment = left_align
        c_name.border = thin_border

        dob_str = student.date_of_birth.strftime("%d/%m/%Y") if student.date_of_birth else ""
        c_dob = ws.cell(row=current_row, column=4, value=dob_str)
        c_dob.font = regular_style
        c_dob.alignment = center_align
        c_dob.border = thin_border

        attendance_map = {}
        if fill_attendance and not is_blank_sheet:
            stmt_att = (
                select(Attendance)
                .where(
                    Attendance.student_id == student.id,
                    Attendance.class_id == class_id,
                    Attendance.date.in_(session_dates)
                )
            )
            att_records = db.scalars(stmt_att).all()
            attendance_map = {r.date: r.status for r in att_records}
            
        present_count = 0
        for i in range(num_cols):
            col_idx = 5 + i
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = regular_style
            cell.alignment = center_align
            cell.border = thin_border
            
            if not is_blank_sheet:
                d = session_dates[i]
                status = attendance_map.get(d)
                if status in ["P", "M"]:
                    cell.value = "X"
                    present_count += 1
                elif status == "V":
                    cell.value = ""
                else:
                    cell.value = ""
            else:
                cell.value = ""
                
        cell_gc = ws.cell(row=current_row, column=total_cols)
        cell_gc.font = regular_style
        cell_gc.border = thin_border
        
        if fill_attendance and not is_blank_sheet:
            cell_gc.value = f"Đủ: {present_count}/{len(session_dates)}"
            cell_gc.alignment = center_align
        else:
            cell_gc.value = ""
            cell_gc.alignment = left_align
            
        current_row += 1
        
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 13
    for i in range(num_cols):
        col_letter = get_column_letter(5 + i)
        ws.column_dimensions[col_letter].width = 8
    ws.column_dimensions[col_letter_last].width = 16
    
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_tuition_excel(db: Session, month: int, year: int, class_id: int | None = None, status: str | None = None) -> bytes:
    from app.services.tuition_service import list_records
    
    records = list_records(db, month, year, class_id)
    
    filtered_records = []
    for row in records:
        prior_debt_stmt = (
            select(func.coalesce(func.sum(TuitionRecord.total_amount - TuitionRecord.paid_amount), 0))
            .where(
                TuitionRecord.student_id == row.student_id,
                (TuitionRecord.year < row.year) | ((TuitionRecord.year == row.year) & (TuitionRecord.month < row.month))
            )
        )
        row.prior_debt = db.scalar(prior_debt_stmt) or 0
        row.grand_total = row.total_amount + row.prior_debt
        row.remaining_debt = max(0, row.total_amount - row.paid_amount)
        
        stat = row.payment_status or "unpaid"
        if status == "da_thu" and stat not in ["paid", "overpaid"]:
            continue
        if status == "chua_thu" and stat in ["paid", "overpaid"]:
            continue
            
        filtered_records.append(row)
        
    wb = Workbook()
    ws = wb.active
    ws.title = f"Học phí T{month:02d}-{year}"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="0F2A33")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True, color="10202B")
    regular_font = Font(name=font_family, size=11, color="10202B")
    
    fill_header = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="CBD5DC"),
        right=Side(style="thin", color="CBD5DC"),
        top=Side(style="thin", color="CBD5DC"),
        bottom=Side(style="thin", color="CBD5DC")
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Title Block
    ws.merge_cells("A1:K1")
    ws["A1"] = f"BẢNG DANH SÁCH THU HỌC PHÍ - THÁNG {month:02d}/{year}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Table Headers
    headers = [
        "STT", "Mã học sinh", "Họ tên học sinh", "Lớp học", 
        "Số buổi", "Học phí kỳ này", "Nợ cũ", "Tổng cần đóng", 
        "Đã đóng", "Còn nợ", "Trạng thái"
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 28
    
    current_row = 4
    for idx, r in enumerate(filtered_records, start=1):
        ws.row_dimensions[current_row].height = 22
        row_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
        
        cells = [
            ws.cell(row=current_row, column=1, value=idx),
            ws.cell(row=current_row, column=2, value=r.student.student_code),
            ws.cell(row=current_row, column=3, value=r.student.full_name),
            ws.cell(row=current_row, column=4, value=", ".join(item.class_name for item in r.items)),
            ws.cell(row=current_row, column=5, value=r.total_sessions),
            ws.cell(row=current_row, column=6, value=r.total_amount),
            ws.cell(row=current_row, column=7, value=r.prior_debt),
            ws.cell(row=current_row, column=8, value=r.grand_total),
            ws.cell(row=current_row, column=9, value=r.paid_amount),
            ws.cell(row=current_row, column=10, value=r.remaining_debt),
            ws.cell(row=current_row, column=11, value="Đã tất toán" if r.payment_status in ["paid", "overpaid"] else "Chưa tất toán" if r.payment_status == "unpaid" else "Mới đóng một phần")
        ]
        
        for c_idx, c in enumerate(cells, start=1):
            c.font = regular_font
            c.border = thin_border
            if row_fill.fill_type:
                c.fill = row_fill
                
            if c_idx in [1, 2, 5, 11]:
                c.alignment = center_align
            elif c_idx in [3, 4]:
                c.alignment = left_align
            else:
                c.alignment = right_align
                c.number_format = "#,##0"
                
        current_row += 1
        
    col_widths = {
        "A": 6, "B": 15, "C": 25, "D": 30, "E": 10, 
        "F": 15, "G": 15, "H": 15, "I": 15, "J": 15, "K": 20
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
        
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_payroll_excel(db: Session, month: int, year: int, status: str | None = None) -> bytes:
    from app.models import TeacherSalaryRecord
    
    stmt = select(TeacherSalaryRecord).options(
        selectinload(TeacherSalaryRecord.teacher),
        selectinload(TeacherSalaryRecord.items)
    )
    if month:
        stmt = stmt.where(TeacherSalaryRecord.month == month)
    if year:
        stmt = stmt.where(TeacherSalaryRecord.year == year)
        
    rows = db.scalars(stmt.order_by(TeacherSalaryRecord.year.desc(), TeacherSalaryRecord.month.desc())).all()
    
    filtered_records = []
    for row in rows:
        prior_unpaid_stmt = (
            select(func.coalesce(func.sum(TeacherSalaryRecord.total_amount - TeacherSalaryRecord.paid_amount), 0))
            .where(
                TeacherSalaryRecord.teacher_id == row.teacher_id,
                (TeacherSalaryRecord.year < row.year) | ((TeacherSalaryRecord.year == row.year) & (TeacherSalaryRecord.month < row.month))
            )
        )
        row.prior_unpaid = db.scalar(prior_unpaid_stmt) or 0
        row.grand_total = row.total_amount + row.prior_unpaid
        row.remaining = max(0, row.total_amount - row.paid_amount)
        
        stat = row.payment_status or "unpaid"
        if status == "da_chi" and stat not in ["paid", "overpaid"]:
            continue
        if status == "chua_chi" and stat in ["paid", "overpaid"]:
            continue
            
        filtered_records.append(row)
        
    wb = Workbook()
    ws = wb.active
    ws.title = f"Lương GV T{month:02d}-{year}"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="0F2A33")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True, color="10202B")
    regular_font = Font(name=font_family, size=11, color="10202B")
    
    fill_header = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="CBD5DC"),
        right=Side(style="thin", color="CBD5DC"),
        top=Side(style="thin", color="CBD5DC"),
        bottom=Side(style="thin", color="CBD5DC")
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Title Block
    ws.merge_cells("A1:J1")
    ws["A1"] = f"BẢNG THANH TOÁN LƯƠNG GIÁO VIÊN - THÁNG {month:02d}/{year}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Table Headers
    headers = [
        "STT", "Họ tên giáo viên", "Lớp dạy", "Tổng số buổi dạy",
        "Lương kỳ này", "Nợ lương cũ", "Tổng cần chi", 
        "Đã chi", "Còn nợ lương", "Trạng thái"
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 28
    
    current_row = 4
    for idx, r in enumerate(filtered_records, start=1):
        ws.row_dimensions[current_row].height = 22
        row_fill = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
        
        total_sessions = sum(item.sessions_count for item in r.items)
        
        cells = [
            ws.cell(row=current_row, column=1, value=idx),
            ws.cell(row=current_row, column=2, value=r.teacher.full_name),
            ws.cell(row=current_row, column=3, value=", ".join(item.class_name for item in r.items)),
            ws.cell(row=current_row, column=4, value=total_sessions),
            ws.cell(row=current_row, column=5, value=r.total_amount),
            ws.cell(row=current_row, column=6, value=r.prior_unpaid),
            ws.cell(row=current_row, column=7, value=r.grand_total),
            ws.cell(row=current_row, column=8, value=r.paid_amount),
            ws.cell(row=current_row, column=9, value=r.remaining),
            ws.cell(row=current_row, column=10, value="Đã tất toán" if r.payment_status in ["paid", "overpaid"] else "Chưa tất toán" if r.payment_status == "unpaid" else "Chi trả một phần")
        ]
        
        for c_idx, c in enumerate(cells, start=1):
            c.font = regular_font
            c.border = thin_border
            if row_fill.fill_type:
                c.fill = row_fill
                
            if c_idx in [1, 4, 10]:
                c.alignment = center_align
            elif c_idx in [2, 3]:
                c.alignment = left_align
            else:
                c.alignment = right_align
                c.number_format = "#,##0"
                
        current_row += 1
        
    col_widths = {
        "A": 6, "B": 25, "C": 30, "D": 15, "E": 15, 
        "F": 15, "G": 15, "H": 15, "I": 15, "J": 20
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
        
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
